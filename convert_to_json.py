"""
入試データベース.xlsx → data/data.json 変換 (GitHub Pages 用)
- 統合DB (大学+学部+学科+方式ごとに 過去3年倍率 + 共テボーダー + 必要科目)
- 共テ科目別平均 (年度×科目)
- 大学リスト + 学部学科リスト (五十音順)
"""
import json
from pathlib import Path
from collections import defaultdict, OrderedDict
from openpyxl import load_workbook

BASE = Path(r"C:\Users\kogat\Documents\Claude")
XLSX = BASE / "入試データベース.xlsx"
OUT  = BASE / "jukenyosoku-web" / "data" / "data.json"

def num(v):
    if v is None: return None
    if isinstance(v, float) and v.is_integer(): return int(v)
    return v

def main():
    wb = load_workbook(XLSX, data_only=True)

    # ── 統合DB ──
    dbws = wb["統合DB"]
    headers = [c.value for c in dbws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    records = []
    for r in dbws.iter_rows(min_row=2, values_only=True):
        if not r[idx["大学"]]: continue
        rec = {
            "u": r[idx["大学"]],
            "f": r[idx["学部"]] or "",
            "d": r[idx["学科"]] or "",
            "m": r[idx["方式"]],
            "kind": r[idx["大学種別"]] or "",
            "bairitsu": {},
            "border": num(r[idx["ボーダー得点率"]]),
            "hensa": num(r[idx["ボーダー偏差値"]]),
            "kyo_man": num(r[idx["共テ満点"]]),
            "ko_man": num(r[idx["2次満点"]]),
            "ratio": num(r[idx["個別配点比率"]]),
            "kyo_subj": r[idx["共テ科目"]] or "",
            "ko_subj": r[idx["2次科目"]] or "",
        }
        for y in (2025, 2024, 2023):
            bo = num(r[idx[f"募集_{y}"]])
            shi = num(r[idx[f"志願_{y}"]])
            ju = num(r[idx[f"受験_{y}"]])
            go = num(r[idx[f"合格_{y}"]])
            ku = num(r[idx[f"競争率_{y}"]])
            kuprv = num(r[idx[f"前年比競争率_in_{y}"]])
            if any(v is not None for v in (bo, shi, ju, go, ku)):
                rec["bairitsu"][str(y)] = {
                    "boshu": bo, "shigan": shi, "juken": ju,
                    "gokaku": go, "ku": ku, "kuprv": kuprv,
                }
        records.append(rec)

    # ── 大学リスト + 学部学科リスト (リストシートから五十音順) ──
    lws = wb["リスト"]
    universities = []
    for r in range(2, lws.max_row + 1):
        v = lws.cell(r, 1).value
        if v: universities.append(v)

    # 各大学の学部学科リスト (リストシートの列ヘッダから)
    departments = OrderedDict()
    for c in range(3, lws.max_column + 1):
        head = lws.cell(1, c).value
        if not head: continue
        depts = []
        for r in range(2, lws.max_row + 1):
            v = lws.cell(r, c).value
            if v: depts.append(v)
        departments[head] = depts

    # 共テ検索用の学部学科リスト (リスト_共テ)
    lkws = wb["リスト_共テ"]
    departments_kyo = OrderedDict()
    for c in range(3, lkws.max_column + 1):
        head = lkws.cell(1, c).value
        if not head: continue
        depts = []
        for r in range(2, lkws.max_row + 1):
            v = lkws.cell(r, c).value
            if v: depts.append(v)
        departments_kyo[head] = depts

    # ── 共テ科目別平均 ──
    aws = wb["共テ科目別平均"]
    kyotest_avg = defaultdict(dict)  # year → {subject: avg}
    for r in aws.iter_rows(min_row=2, values_only=True):
        year, katei, kubun, subject, ninzu, heikin, manten, biko = r[0:8]
        if year is None or not subject: continue
        kyotest_avg[str(year)][subject] = {
            "avg": num(heikin), "man": num(manten),
        }

    data = {
        "meta": {"records": len(records), "universities": len(universities)},
        "universities": universities,
        "departments": departments,
        "departments_kyo": departments_kyo,
        "records": records,
        "kyotest_avg": kyotest_avg,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))

    size_kb = OUT.stat().st_size / 1024
    print(f"data.json 出力完了: {OUT}")
    print(f"  records: {len(records)}")
    print(f"  universities: {len(universities)}")
    print(f"  共テ平均 年度: {list(kyotest_avg.keys())}")
    print(f"  ファイルサイズ: {size_kb:.0f} KB")

if __name__ == "__main__":
    main()
